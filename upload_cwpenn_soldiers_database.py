from __future__ import annotations

import hashlib
import json
import os
import re
import sys
import time
from datetime import datetime
from pathlib import Path
from typing import Any

import psycopg2
from dotenv import load_dotenv
from openpyxl import load_workbook
from psycopg2 import sql


SCRIPT_DIR = Path(__file__).resolve().parent
ENV_PATH = SCRIPT_DIR / ".env"
if not ENV_PATH.exists():
    ENV_PATH = SCRIPT_DIR / "backend" / ".env"
load_dotenv(ENV_PATH)

SOURCE_DB_DSN = os.getenv("SOURCE_DB_CONNECTION_STRING") or os.getenv("DB_CONNECTION_STRING", "")
BACKUP_DB_DSN = os.getenv("BACKUP_DB_CONNECTION_STRING", "")
AUDIT_USERNAME = os.getenv("AUDIT_USERNAME", "cwpenn_import")

INPUT_FILE = SCRIPT_DIR / "camp_william_penn_civil_war_troops_v4.xlsx"
SHEET_NAME = "Sheet1"
CHECKPOINT_FILE = SCRIPT_DIR / "upload_cwpenn_soldiers_database_checkpoint.json"
LOG_FILE = SCRIPT_DIR / "upload_cwpenn_soldiers_database.log"


if not SOURCE_DB_DSN:
    print("ERROR: SOURCE_DB_CONNECTION_STRING (or DB_CONNECTION_STRING) is not set in .env")
    sys.exit(1)

if not BACKUP_DB_DSN:
    print("ERROR: BACKUP_DB_CONNECTION_STRING is not set in .env")
    sys.exit(1)

if SOURCE_DB_DSN == BACKUP_DB_DSN:
    print("ERROR: Source and backup connection strings must be different.")
    sys.exit(1)


def log(message: str) -> None:
    ts = datetime.now().strftime("%H:%M:%S")
    line = f"[{ts}] {message}"
    print(line, flush=True)
    LOG_FILE.parent.mkdir(parents=True, exist_ok=True)
    with LOG_FILE.open("a", encoding="utf-8") as handle:
        handle.write(line + "\n")


def is_blank_like(value: Any) -> bool:
    if value is None:
        return True

    text = str(value).strip()
    return text.lower() in {
        "",
        "nan",
        "nat",
        "none",
        "null",
        "-",
        "--",
        "----",
        "[blank]",
        "n/a",
        "na",
    }


def clean_text(value: Any) -> str | None:
    if is_blank_like(value):
        return None
    return str(value).strip()


def raw_text(value: Any) -> str | None:
    if value is None:
        return None
    return str(value).strip()


def pick_cell(row: dict[str, Any], *keys: str) -> str | None:
    for key in keys:
        value = clean_text(row.get(key))
        if value is not None:
            return value
    return None


def parse_int(value: Any) -> int | None:
    text = clean_text(value)
    if text is None:
        return None
    match = re.search(r"-?\d+", text)
    if not match:
        return None
    try:
        return int(match.group(0))
    except ValueError:
        return None


def normalize_column_name(name: str) -> str:
    normalized = re.sub(r"\s+", "_", name.strip().lower())
    normalized = re.sub(r"[^a-z0-9_]+", "_", normalized)
    normalized = re.sub(r"_+", "_", normalized).strip("_")
    if not normalized:
        normalized = "column"

    if len(normalized) > 50:
        digest = hashlib.sha1(name.encode("utf-8")).hexdigest()[:8]
        normalized = f"{normalized[:50].rstrip('_')}_{digest}"

    return normalized[:63]


def make_unique_column_names(headers: list[str]) -> list[str]:
    seen: dict[str, int] = {}
    unique_names: list[str] = []

    for header in headers:
        base = normalize_column_name(header)
        count = seen.get(base, 0)
        seen[base] = count + 1
        if count == 0:
            unique_names.append(base)
            continue

        suffix = f"_{count + 1}"
        trimmed = base[: 63 - len(suffix)].rstrip("_")
        unique_names.append(f"{trimmed}{suffix}")

    return unique_names


def parse_name_from_folder(folder_name: str | None) -> tuple[str | None, str | None]:
    if not folder_name:
        return None, None

    cleaned = folder_name.strip()
    cleaned = re.sub(r"\s+\d+(?:st|nd|rd|th)\s+U\.?S\.?C\.?T\.?.*$", "", cleaned, flags=re.I)

    if "," in cleaned:
        surname, remainder = cleaned.split(",", 1)
        remainder = re.split(
            r"\s+(?:Co\b|Pvt\b|Sgt\b|Cpl\b|Corpl\b|Corp\b|Mus\b|Drummer\b|Bugler\b|Capt\b|Lieut\b|Major\b|Gen\b|General\b|Private\b|Sergeant\b)",
            remainder,
            maxsplit=1,
            flags=re.I,
        )[0]
        first_name = remainder.strip()
        surname = surname.strip()
        return first_name or None, surname or None

    parts = cleaned.split()
    if len(parts) == 1:
        return None, parts[0]
    return " ".join(parts[:-1]).strip() or None, parts[-1].strip() or None


def derive_core_name(row: dict[str, Any]) -> tuple[str | None, str | None]:
    first_name = clean_text(row.get("FirstName"))
    last_name = clean_text(row.get("Surname"))

    if first_name and last_name:
        return first_name, last_name

    parsed_first, parsed_last = parse_name_from_folder(clean_text(row.get("Soldier_Folder_Name")))
    if not first_name:
        first_name = parsed_first
    if not last_name:
        last_name = parsed_last

    return first_name, last_name


def derive_birth_date(row: dict[str, Any]) -> str | None:
    estimated = clean_text(row.get("Estimated_Birth_Year"))
    if estimated:
        range_match = re.match(r"^(\d{4})\s*-\s*(\d{4})$", estimated)
        if range_match:
            return f"{range_match.group(1)}-01-01"

        match = re.match(r"^(\d{4})", estimated)
        if match:
            return f"{match.group(1)}-01-01"

    return None


def derive_location_fields(row: dict[str, Any]) -> dict[str, str | None] | None:
    city = clean_text(row.get("City"))
    county = clean_text(row.get("County"))
    state = clean_text(row.get("Birth_State_Country"))
    if not state:
        raw_state = clean_text(row.get("State"))
        if raw_state and raw_state not in {"United States", "USCT"}:
            state = raw_state

    coordinates = clean_text(row.get("Coordinates"))
    country = clean_text(row.get("Country")) or clean_text(row.get("Served_for"))
    landmark = clean_text(row.get("Military_Unit"))

    if not any([city, county, state, coordinates]):
        return None

    return {
        "city": city,
        "county": county,
        "state": state,
        "coordinates": coordinates,
        "country": country,
        "landmark": landmark,
    }


def build_source_row(headers: list[str], values: tuple[Any, ...]) -> dict[str, Any]:
    return {header: values[idx] for idx, header in enumerate(headers)}


def format_duration(seconds: float) -> str:
    seconds = max(0, int(round(seconds)))
    hours, remainder = divmod(seconds, 3600)
    minutes, secs = divmod(remainder, 60)
    if hours:
        return f"{hours:02d}:{minutes:02d}:{secs:02d}"
    return f"{minutes:02d}:{secs:02d}"


def load_checkpoint() -> dict[str, Any]:
    if not CHECKPOINT_FILE.exists():
        return {}
    try:
        return json.loads(CHECKPOINT_FILE.read_text(encoding="utf-8"))
    except json.JSONDecodeError:
        return {}


def save_checkpoint(*, row_number: int, source_id: int | None, processed: int, total: int) -> None:
    payload = {
        "row_number": row_number,
        "source_id": source_id,
        "processed": processed,
        "total": total,
        "updated_at": datetime.now().isoformat(),
    }
    CHECKPOINT_FILE.write_text(json.dumps(payload, indent=2), encoding="utf-8")


def compute_eta(elapsed_seconds: float, processed: int, total: int) -> float | None:
    if processed <= 0:
        return None
    avg = elapsed_seconds / processed
    remaining = total - processed
    return avg * remaining


def ensure_audit_objects(cur) -> None:
    cur.execute("CREATE SCHEMA IF NOT EXISTS audit;")
    cur.execute(
        """
        CREATE INDEX IF NOT EXISTS idx_audit_table_time
            ON audit.logged_actions (table_name, changed_at DESC);
        """
    )
    cur.execute(
        """
        CREATE INDEX IF NOT EXISTS idx_audit_changed_by
            ON audit.logged_actions (changed_by);
        """
    )
    cur.execute(
        """
        CREATE OR REPLACE FUNCTION audit.log_table_changes()
        RETURNS TRIGGER
        LANGUAGE plpgsql
        SECURITY DEFINER
        SET search_path = audit, public
        AS $$
        DECLARE
            v_row_id        TEXT;
            v_changed_cols  TEXT[];
            v_old_json      JSONB;
            v_new_json      JSONB;
            v_action        TEXT := TG_OP;
            v_changed_by    TEXT;
            v_pk_col        TEXT := TG_ARGV[0];
        BEGIN
            v_changed_by := NULLIF(current_setting('audit.username', true), '');

            IF TG_OP IN ('UPDATE', 'DELETE') THEN
                v_row_id := row_to_json(OLD)->>v_pk_col;
            ELSE
                v_row_id := row_to_json(NEW)->>v_pk_col;
            END IF;

            IF TG_OP = 'INSERT' THEN
                v_new_json     := row_to_json(NEW);
                v_changed_cols := ARRAY(SELECT jsonb_object_keys(v_new_json));
            ELSIF TG_OP = 'UPDATE' THEN
                v_old_json := row_to_json(OLD);
                v_new_json := row_to_json(NEW);

                SELECT ARRAY_AGG(key)
                INTO v_changed_cols
                FROM jsonb_object_keys(v_new_json) AS t(key)
                WHERE v_old_json->>key IS DISTINCT FROM v_new_json->>key;

                IF v_changed_cols IS NULL OR array_length(v_changed_cols, 1) = 0 THEN
                    RETURN NEW;
                END IF;
            ELSIF TG_OP = 'DELETE' THEN
                v_old_json     := row_to_json(OLD);
                v_changed_cols := ARRAY(SELECT jsonb_object_keys(v_old_json));
            END IF;

            INSERT INTO audit.logged_actions (
                schema_name, table_name, action, row_id,
                changed_by, old_values, new_values, changed_columns
            ) VALUES (
                TG_TABLE_SCHEMA, TG_TABLE_NAME, v_action, v_row_id,
                v_changed_by, v_old_json, v_new_json, v_changed_cols
            );

            RETURN COALESCE(NEW, OLD);
        END;
        $$;
        """
    )


def ensure_locations_table(cur) -> None:
    cur.execute(
        """
        CREATE TABLE IF NOT EXISTS public.locations (
            locations_id BIGSERIAL PRIMARY KEY,
            created_at TIMESTAMPTZ NOT NULL DEFAULT now()
        );
        """
    )
    for column_name, ddl in [
        ("city", "TEXT"),
        ("county", "TEXT"),
        ("state", "TEXT"),
        ("coordinates", "TEXT"),
        ("country", "TEXT"),
        ("landmark", "TEXT"),
    ]:
        cur.execute(
            sql.SQL("ALTER TABLE public.locations ADD COLUMN IF NOT EXISTS {} {};").format(
                sql.Identifier(column_name),
                sql.SQL(ddl),
            )
        )
    cur.execute(
        """
        CREATE INDEX IF NOT EXISTS idx_locations_dedupe
            ON public.locations (city, county, state, country, landmark);
        """
    )


def ensure_family_members_table(cur) -> None:
    cur.execute(
        """
        CREATE TABLE IF NOT EXISTS public.family_members (
            member_id BIGSERIAL PRIMARY KEY,
            created_at TIMESTAMPTZ NOT NULL DEFAULT now()
        );
        """
    )
    columns = [
        ("directory_id", "BIGINT"),
        ("first_name", "TEXT"),
        ("last_name", "TEXT"),
        ("alias", "TEXT"),
        ("gender", "TEXT"),
        ("race", "TEXT"),
        ("ethnicity", "TEXT"),
        ("generation_number", "INTEGER"),
        ("father_id", "BIGINT"),
        ("mother_id", "BIGINT"),
        ("birth_date", "TEXT"),
        ("birth_location_id", "BIGINT"),
        ("death_date", "TEXT"),
        ("death_location_id", "BIGINT"),
        ("marriage_date", "TEXT"),
        ("marriage_location_id", "BIGINT"),
        ("military_service", "TEXT"),
        ("branch", "TEXT"),
        ("war", "TEXT"),
        ("spouse_id", "BIGINT"),
    ]
    for column_name, ddl in columns:
        cur.execute(
            sql.SQL("ALTER TABLE public.family_members ADD COLUMN IF NOT EXISTS {} {};").format(
                sql.Identifier(column_name),
                sql.SQL(ddl),
            )
        )
    cur.execute(
        """
        CREATE UNIQUE INDEX IF NOT EXISTS idx_family_members_directory_id
            ON public.family_members (directory_id);
        """
    )


def ensure_camp_william_penn_table(cur, camp_columns: list[str]) -> None:
    cur.execute(
        """
        CREATE TABLE IF NOT EXISTS public.camp_william_penn (
            camp_william_penn_id BIGSERIAL PRIMARY KEY,
            member_id BIGINT,
            source_id BIGINT,
            created_at TIMESTAMPTZ NOT NULL DEFAULT now()
        );
        """
    )
    for column_name in camp_columns:
        cur.execute(
            sql.SQL("ALTER TABLE public.camp_william_penn ADD COLUMN IF NOT EXISTS {} TEXT;").format(
                sql.Identifier(column_name)
            )
        )
    cur.execute(
        """
        CREATE UNIQUE INDEX IF NOT EXISTS idx_cwp_source_id
            ON public.camp_william_penn (source_id);
        """
    )
    cur.execute(
        """
        CREATE UNIQUE INDEX IF NOT EXISTS idx_cwp_member_id
            ON public.camp_william_penn (member_id);
        """
    )


def attach_audit_trigger(cur, table_name: str, pk_column: str) -> None:
    trigger_name = f"trg_audit_{table_name}"
    cur.execute(
        sql.SQL("DROP TRIGGER IF EXISTS {} ON public.{};").format(
            sql.Identifier(trigger_name),
            sql.Identifier(table_name),
        )
    )
    cur.execute(
        sql.SQL(
            """
            CREATE TRIGGER {}
                AFTER INSERT OR UPDATE OR DELETE
                ON public.{}
                FOR EACH ROW
                EXECUTE FUNCTION audit.log_table_changes({});
            """
        ).format(
            sql.Identifier(trigger_name),
            sql.Identifier(table_name),
            sql.Literal(pk_column),
        )
    )


def ensure_database_objects(cur, camp_columns: list[str]) -> None:
    ensure_audit_objects(cur)
    ensure_locations_table(cur)
    ensure_family_members_table(cur)
    ensure_camp_william_penn_table(cur, camp_columns)
    attach_audit_trigger(cur, "camp_william_penn", "camp_william_penn_id")


def set_audit_username(cur, username: str) -> None:
    cur.execute("SELECT set_config('audit.username', %s, false);", (username,))


def get_next_member_id(cur) -> int:
    cur.execute("SELECT COALESCE(MAX(member_id), 0) + 1 FROM public.family_members;")
    return int(cur.fetchone()[0])


def sync_sequence_to_max(cur, table_name: str, column_name: str) -> None:
    cur.execute(
        sql.SQL("SELECT pg_get_serial_sequence(%s, %s);"),
        (table_name, column_name),
    )
    sequence_name = cur.fetchone()[0]
    if not sequence_name:
        return

    cur.execute(
        sql.SQL("SELECT COALESCE(MAX({column}), 0) FROM {table};").format(
            column=sql.Identifier(column_name),
            table=sql.SQL(table_name),
        )
    )
    max_value = int(cur.fetchone()[0])
    if max_value <= 0:
        cur.execute(
            sql.SQL("SELECT setval(%s, 1, false);"),
            (sequence_name,),
        )
        return

    cur.execute(
        sql.SQL("SELECT setval(%s, %s, true);"),
        (sequence_name, max_value),
    )


def get_or_create_location(cur, location: dict[str, str | None] | None) -> int | None:
    if not location:
        return None

    cur.execute(
        """
        SELECT locations_id
        FROM public.locations
        WHERE city IS NOT DISTINCT FROM %s
          AND county IS NOT DISTINCT FROM %s
          AND state IS NOT DISTINCT FROM %s
          AND country IS NOT DISTINCT FROM %s
          AND landmark IS NOT DISTINCT FROM %s
        ORDER BY locations_id
        LIMIT 1
        """,
        (
            location.get("city"),
            location.get("county"),
            location.get("state"),
            location.get("country"),
            location.get("landmark"),
        ),
    )
    existing = cur.fetchone()
    if existing:
        location_id = existing[0]
        if location.get("coordinates"):
            cur.execute(
                """
                UPDATE public.locations
                SET coordinates = %s
                WHERE locations_id = %s
                  AND (coordinates IS NULL OR coordinates = '')
                """,
                (location.get("coordinates"), location_id),
            )
        return location_id

    cur.execute(
        """
        INSERT INTO public.locations (city, county, state, coordinates, country, landmark)
        VALUES (%s, %s, %s, %s, %s, %s)
        RETURNING locations_id
        """,
        (
            location.get("city"),
            location.get("county"),
            location.get("state"),
            location.get("coordinates"),
            location.get("country"),
            location.get("landmark"),
        ),
    )
    return cur.fetchone()[0]


def build_family_member_payload(
    row: dict[str, Any],
    location_id: int | None,
) -> dict[str, Any]:
    first_name = pick_cell(row, "First_Name", "FirstName")
    last_name = pick_cell(row, "Surname", "Last_Name", "LastName")
    if not first_name or not last_name:
        fallback_first, fallback_last = derive_core_name(row)
        first_name = first_name or fallback_first
        last_name = last_name or fallback_last

    birth_date = derive_birth_date(row)

    return {
        "directory_id": parse_int(row.get("Id")),
        "first_name": first_name,
        "last_name": last_name,
        "alias": "-",
        "gender": pick_cell(row, "Gender"),
        "race": pick_cell(row, "Race"),
        "ethnicity": pick_cell(row, "Ethnicity"),
        "generation_number": 1,
        "father_id": None,
        "mother_id": None,
        "birth_date": birth_date,
        "birth_location_id": location_id,
        "death_date": None,
        "death_location_id": None,
        "marriage_date": None,
        "marriage_location_id": None,
        "military_service": clean_text(row.get("Military_Unit")),
        "branch": clean_text(row.get("Branch")),
        "war": clean_text(row.get("Conflict_Period")),
        "spouse_id": None,
    }


def insert_family_member(cur, payload: dict[str, Any], member_id: int) -> int:
    columns = [
        "member_id",
        "directory_id",
        "first_name",
        "last_name",
        "alias",
        "gender",
        "race",
        "ethnicity",
        "generation_number",
        "father_id",
        "mother_id",
        "birth_date",
        "birth_location_id",
        "death_date",
        "death_location_id",
        "marriage_date",
        "marriage_location_id",
        "military_service",
        "branch",
        "war",
        "spouse_id",
    ]
    query = sql.SQL(
        """
        INSERT INTO public.family_members ({fields})
        VALUES ({values})
        RETURNING member_id
        """
    ).format(
        fields=sql.SQL(", ").join(map(sql.Identifier, columns)),
        values=sql.SQL(", ").join(sql.Placeholder() for _ in columns),
    )
    values = [member_id] + [payload[column] for column in columns[1:]]
    cur.execute(query, values)
    return cur.fetchone()[0]


def build_camp_payload(
    source_id: int,
    member_id: int,
    source_row: dict[str, Any],
    camp_field_specs: list[tuple[int, str]],
) -> dict[str, Any]:
    payload: dict[str, Any] = {
        "source_id": source_id,
        "member_id": member_id,
    }
    for index, db_column in camp_field_specs:
        payload[db_column] = raw_text(source_row["values"][index])
    return payload


def upsert_camp_record(cur, payload: dict[str, Any], camp_columns: list[str]) -> int:
    columns = ["source_id", "member_id"] + camp_columns
    update_columns = [column for column in columns if column != "source_id"]
    query = sql.SQL(
        """
        INSERT INTO public.camp_william_penn ({fields})
        VALUES ({values})
        ON CONFLICT (source_id) DO UPDATE SET
            {updates}
        RETURNING camp_william_penn_id
        """
    ).format(
        fields=sql.SQL(", ").join(map(sql.Identifier, columns)),
        values=sql.SQL(", ").join(sql.Placeholder() for _ in columns),
        updates=sql.SQL(", ").join(
            sql.SQL("{} = EXCLUDED.{}").format(sql.Identifier(column), sql.Identifier(column))
            for column in update_columns
        ),
    )
    cur.execute(query, [payload[column] for column in columns])
    return cur.fetchone()[0]


def load_workbook_rows(path: Path, sheet_name: str) -> tuple[list[str], list[tuple[Any, ...]]]:
    workbook = load_workbook(path, read_only=True, data_only=True)
    if sheet_name not in workbook.sheetnames:
        raise ValueError(f"Worksheet not found: {sheet_name}")

    worksheet = workbook[sheet_name]
    rows = list(worksheet.iter_rows(values_only=True))
    if not rows:
        return [], []

    headers = [str(value).strip() if value is not None else "" for value in rows[0]]
    data_rows = [tuple(row) for row in rows[1:]]
    return headers, data_rows


def build_camp_field_specs(headers: list[str]) -> list[tuple[int, str]]:
    specs: list[tuple[int, str]] = []
    seen: dict[str, int] = {}
    for index, header in enumerate(headers):
        if header == "Id":
            continue
        base = normalize_column_name(header)
        count = seen.get(base, 0)
        seen[base] = count + 1
        if count == 0:
            db_name = base
        else:
            suffix = f"_{count + 1}"
            trimmed = base[: 63 - len(suffix)].rstrip("_")
            db_name = f"{trimmed}{suffix}"
        specs.append((index, db_name))
    if not any(db_name == "estimated_birth_year" for _, db_name in specs):
        estimated_index = headers.index("Estimated_Birth_Year") if "Estimated_Birth_Year" in headers else None
        if estimated_index is not None:
            specs.append((estimated_index, "estimated_birth_year"))
    return specs


def process_row(
    cur,
    source_row: dict[str, Any],
    camp_field_specs: list[tuple[int, str]],
    member_state: dict[str, int],
) -> tuple[int, int]:
    source_id = parse_int(source_row.get("Id"))
    if source_id is None:
        raise ValueError("Row is missing a valid Id value.")

    location_id = get_or_create_location(cur, derive_location_fields(source_row))
    family_payload = build_family_member_payload(source_row, location_id)
    member_id = member_state["next_member_id"]
    inserted_member_id = insert_family_member(cur, family_payload, member_id)
    if inserted_member_id != member_id:
        member_id = inserted_member_id
    member_state["next_member_id"] = member_id + 1

    camp_payload = build_camp_payload(
        source_id=source_id,
        member_id=member_id,
        source_row=source_row,
        camp_field_specs=camp_field_specs,
    )
    upsert_camp_record(cur, camp_payload, [db_name for _, db_name in camp_field_specs])
    return source_id, member_id


def print_mapping_summary(headers: list[str], camp_field_specs: list[tuple[int, str]]) -> None:
    family_member_mapping = [
        ("First_Name / FirstName", "family_members.first_name"),
        ("Surname / Last_Name / LastName", "family_members.last_name"),
        ("-", "family_members.alias"),
        ("Gender", "family_members.gender"),
        ("Race", "family_members.race"),
        ("Ethnicity", "family_members.ethnicity"),
        ("Estimated_Birth_Year", "family_members.birth_date"),
    ]
    supplemental_columns = [db_name for _, db_name in camp_field_specs]

    log("Field mapping summary:")
    for source_field, target_field in family_member_mapping:
        log(f"  {source_field} -> {target_field}")

    log(f"  All remaining Excel columns -> public.camp_william_penn ({len(supplemental_columns)} columns)")
    preview = ", ".join(supplemental_columns[:12])
    if len(supplemental_columns) > 12:
        preview += ", ..."
    log(f"  Supplemental columns: {preview}")


def main() -> None:
    if not INPUT_FILE.exists():
        raise FileNotFoundError(f"Input workbook not found: {INPUT_FILE}")

    headers, data_rows = load_workbook_rows(INPUT_FILE, SHEET_NAME)
    if not headers or not data_rows:
        raise RuntimeError("No data rows were found in the workbook.")

    camp_field_specs = build_camp_field_specs(headers)
    log(f"Loaded {len(data_rows)} workbook rows from {INPUT_FILE.name}.")
    log(f"Preparing {len(camp_field_specs)} supplemental columns for camp_william_penn.")
    print_mapping_summary(headers, camp_field_specs)

    checkpoint = load_checkpoint()
    resume_after_source_id = parse_int(checkpoint.get("source_id")) if checkpoint else None
    if resume_after_source_id is not None:
        log(f"Checkpoint detected. Will resume after source Id {resume_after_source_id}.")

    source_conn = psycopg2.connect(SOURCE_DB_DSN)
    backup_conn = psycopg2.connect(BACKUP_DB_DSN)
    source_conn.autocommit = False
    backup_conn.autocommit = False

    processed = 0
    skipped = 0
    total_rows = len(data_rows)
    started_at = time.perf_counter()
    seen_resume_marker = resume_after_source_id is None
    try:
        with source_conn.cursor() as source_cur, backup_conn.cursor() as backup_cur:
            ensure_database_objects(source_cur, [db_name for _, db_name in camp_field_specs])
            ensure_database_objects(backup_cur, [db_name for _, db_name in camp_field_specs])
            for cur in (source_cur, backup_cur):
                sync_sequence_to_max(cur, "audit.logged_actions", "id")
                sync_sequence_to_max(cur, "public.locations", "locations_id")
                sync_sequence_to_max(cur, "public.family_members", "member_id")
                sync_sequence_to_max(cur, "public.camp_william_penn", "camp_william_penn_id")
            set_audit_username(source_cur, AUDIT_USERNAME)
            set_audit_username(backup_cur, AUDIT_USERNAME)

            source_member_state = {"next_member_id": get_next_member_id(source_cur)}
            backup_member_state = {"next_member_id": get_next_member_id(backup_cur)}
            log(
                f"Starting family_members IDs at source={source_member_state['next_member_id']} "
                f"backup={backup_member_state['next_member_id']}"
            )
            if source_member_state["next_member_id"] != backup_member_state["next_member_id"]:
                log("Warning: source and backup member_id starting points differ.")

            source_conn.commit()
            backup_conn.commit()

            for row_number, row_values in enumerate(data_rows, start=2):
                source_row = build_source_row(headers, row_values)
                source_row["values"] = row_values
                source_id = parse_int(source_row.get("Id"))
                display_name = clean_text(source_row.get("Soldier_Folder_Name")) or "unknown record"

                if not seen_resume_marker:
                    if source_id == resume_after_source_id:
                        seen_resume_marker = True
                    skipped += 1
                    continue

                try:
                    log(f"Processing {processed + 1}/{total_rows} | row {row_number} | Id={source_id or 'unknown'} | {display_name}")
                    process_row(source_cur, source_row, camp_field_specs, source_member_state)
                    process_row(backup_cur, source_row, camp_field_specs, backup_member_state)
                    source_conn.commit()
                    backup_conn.commit()
                    processed += 1
                    elapsed = time.perf_counter() - started_at
                    eta = compute_eta(elapsed, processed, total_rows)
                    progress = f"{processed}/{total_rows}"
                    eta_text = format_duration(eta) if eta is not None else "calculating..."
                    log(
                        f"Completed {progress} | elapsed {format_duration(elapsed)} | ETA {eta_text}"
                    )
                    save_checkpoint(
                        row_number=row_number,
                        source_id=source_id,
                        processed=processed,
                        total=total_rows,
                    )
                except Exception as exc:
                    source_conn.rollback()
                    backup_conn.rollback()
                    raise RuntimeError(f"Failed while processing Excel row {row_number}: {exc}") from exc

    finally:
        source_conn.close()
        backup_conn.close()

    elapsed = time.perf_counter() - started_at
    log(
        f"Done. Processed {processed} rows into source and backup databases. "
        f"Skipped {skipped} rows while resuming. Elapsed {format_duration(elapsed)}."
    )


if __name__ == "__main__":
    main()
