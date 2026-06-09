from __future__ import annotations

from pathlib import Path

from openpyxl import load_workbook


INPUT_FILE = Path("camp_william_penn_civil_war_troops_v3.xlsx")
OUTPUT_FILE = Path("camp_william_penn_civil_war_troops_v5.xlsx")
REPLACEMENT_VALUE = "-"


def should_replace(value) -> bool:
    if value is None:
        return True

    if isinstance(value, str):
        stripped = value.strip()
        if not stripped:
            return True
        if stripped == "[Blank]":
            return True

    return False


def main() -> None:
    if not INPUT_FILE.exists():
        raise FileNotFoundError(f"Input file not found: {INPUT_FILE}")

    workbook = load_workbook(INPUT_FILE)

    replacements = 0
    for worksheet in workbook.worksheets:
        for row in worksheet.iter_rows():
            for cell in row:
                if should_replace(cell.value):
                    cell.value = REPLACEMENT_VALUE
                    replacements += 1

    workbook.save(OUTPUT_FILE)
    print(f"Saved {OUTPUT_FILE} with {replacements} replacements.")


if __name__ == "__main__":
    main()
